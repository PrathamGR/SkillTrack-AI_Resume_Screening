import streamlit as st
import pandas as pd
import PyPDF2
import json
from typing import Dict, Any, List, Optional 
from qa_system import ResumeQASystem
from utils import groq_generate 
import streamlit.components.v1 as components
import tempfile
import smtplib
from email.message import EmailMessage
import re

# ----------------------------------------------
# Helper Functions
# ----------------------------------------------

def read_resume(uploaded_file) -> str:
    """Extract text from uploaded resume (or job requirements) file."""
    if uploaded_file.type == "application/pdf":
        pdf_reader = PyPDF2.PdfReader(uploaded_file)
        text = " ".join([page.extract_text() for page in pdf_reader.pages if page.extract_text()])
    else:
        text = uploaded_file.getvalue().decode()
    return text

def parse_llm_response(text: str) -> Dict[str, Any]:
    """
    Parse the LLM response with improved work experience handling and certificates section.
    Returns a nested dictionary.
    """
    sections = {
       "Basic Info": {  
            "Name": "",
            "Email": "",
            "Phone": ""
        },
        "Profile Summary": "",
        "Work Experience": [],  
        "Education": "",
        "Technical Skills": "",
        "Projects": [],
        "Certificates": ""  
    }
    
    section_markers = {
        "name:": ("Basic Info", "Name"),
        "email:": ("Basic Info", "Email"),
        "phone:": ("Basic Info", "Phone"),
        "phone number:": ("Basic Info", "Phone"),
        "profile summary:": ("Profile Summary", None),
        "summary:": ("Profile Summary", None),
        "work experience:": ("Work Experience", None),
        "employment:": ("Work Experience", None),
        "education:": ("Education", None),
        "technical skills:": ("Technical Skills", None),
        "skills:": ("Technical Skills", None),
        "projects:": ("Projects", None),
        "certificates:": ("Certificates", None),
        "certifications:": ("Certificates", None)
    }
    
    lines = text.split('\n')
    current_section = None
    current_subsection = None
    section_content = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        line_lower = line.lower()
        
        # Check if this line starts a new section using our markers.
        new_section = None
        new_subsection = None
        for marker, (section, subsection) in section_markers.items():
            if line_lower.startswith(marker):
                new_section = section
                new_subsection = subsection
                content = line[len(marker):].strip()
                break
        
        if new_section:
            # Save previously accumulated content.
            if current_section and section_content:
                if current_section == "Work Experience":
                    sections[current_section] = parse_work_experience('\n'.join(section_content))
                elif current_section == "Projects":
                    sections[current_section] = parse_projects('\n'.join(section_content))
                elif current_section == "Basic Info" and current_subsection:
                    sections[current_section][current_subsection] = '\n'.join(section_content).strip()
                else:
                    sections[current_section] = '\n'.join(section_content).strip()
            
            # Start a new section.
            current_section = new_section
            current_subsection = new_subsection
            section_content = [content] if content else []
        elif current_section:
            section_content.append(line)
    
    # Save final section.
    if current_section and section_content:
        if current_section == "Work Experience":
            sections[current_section] = parse_work_experience('\n'.join(section_content))
        elif current_section == "Projects":
            sections[current_section] = parse_projects('\n'.join(section_content))
        elif current_section == "Basic Info" and current_subsection:
            sections[current_section][current_subsection] = '\n'.join(section_content).strip()
        else:
            sections[current_section] = '\n'.join(section_content).strip()
    
    return sections

def parse_work_experience(text: str) -> List[Dict[str, Any]]:
    """Parse work experience details into structured data with companies and responsibilities."""
    experiences = []
    current_exp = None
    current_responsibilities = []
    
    lines = text.split('\n')
    i = 0
    
    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1
            continue
        
        # Check if this line appears to announce a new company entry.
        if (line.startswith(('•', '-', '*', '○', '·', '►', '▪', '➢')) and 
            ('(' in line and ')' in line)) or \
           (not line.startswith(('•', '-', '*', '○', '·', '►', '▪', '➢')) and 
            '(' in line and ')' in line):
            
            # Save any previous experience.
            if current_exp and current_responsibilities:
                current_exp['responsibilities'] = current_responsibilities
                experiences.append(current_exp)
            
            if line.startswith(('•', '-', '*', '○', '·', '►', '▪', '➢')):
                line = line[1:].strip()
            
            # Parse company info.
            try:
                company_part, date_part = line.split('(', 1)
                date_part = date_part.rstrip(')')
                
                # Split location if present.
                if ',' in date_part:
                    date_info, location = date_part.rsplit(',', 1)
                else:
                    date_info, location = date_part, ""
                
                current_exp = {
                    'company': company_part.strip(),
                    'duration': date_info.strip(),
                    'location': location.strip(),
                    'responsibilities': []
                }
                current_responsibilities = []
            except ValueError:
                # Handle malformed lines.
                current_exp = {
                    'company': line,
                    'duration': '',
                    'location': '',
                    'responsibilities': []
                }
                current_responsibilities = []
        
        # Check for responsibility.
        elif line.startswith(('•', '-', '*', '○', '·', '►', '▪', '➢')) and current_exp:
            resp = line.lstrip('•-*○·►▪➢ ').strip()
            if resp:
                current_responsibilities.append(resp)
        
        i += 1
    
    # Add final experience.
    if current_exp and current_responsibilities:
        current_exp['responsibilities'] = current_responsibilities
        experiences.append(current_exp)
    
    return experiences

def parse_projects(text: str) -> List[Dict[str, Any]]:
    """Parse project details into structured data."""
    projects = []
    current_project = None
    current_details = []
    
    lines = text.split('\n')
    i = 0
    
    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1
            continue
        
        # Check if line starts with bullet and might be a project title.
        if line.startswith(('•', '-', '*', '○', '·', '►', '▪', '➢')):
            clean_line = line.lstrip('•-*○·►▪➢ ').strip()
            
            # Determine if it's a title.
            is_title = True
            if i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if next_line.startswith(('•', '-', '*', '○', '·', '►', '▪', '➢')):
                    if ("technologies:" in clean_line.lower() or 
                        "developed" in clean_line.lower() or 
                        "implemented" in clean_line.lower() or 
                        "built" in clean_line.lower()):
                        is_title = False
            
            if is_title:
                # Save previous project if exists.
                if current_project and current_details:
                    current_project['details'] = current_details
                    projects.append(current_project)
                
                # Start new project.
                current_project = {
                    'title': clean_line,
                    'details': []
                }
                current_details = []
            elif current_project:
                current_details.append(clean_line)
        
        i += 1
    
    # Add final project.
    if current_project and current_details:
        current_project['details'] = current_details
        projects.append(current_project)
    
    return projects

def extract_email(text: str) -> str:
    """Fallback extraction of an email address using regex."""
    match = re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b', text)
    return match.group(0) if match else ""

def extract_phone(text: str) -> str:
    """Fallback extraction of a phone number using regex."""
    match = re.search(r'(\+?\d{1,2}\s?-?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4})', text)
    return match.group(0) if match else ""

def extract_info(resume_text: str) -> Dict[str, Any]:
    """
    Extract structured resume details using an LLM prompt.
    The prompt instructs the model to output the details EXACTLY in the following format with one header per line.
    """
    prompt = f"""
    Please analyze the following resume and extract the information EXACTLY in the following format with one header per line (do not include any extra text):

    Name: [Full Name]
    Email: [Email Address]
    Phone: [Phone Number]
    Profile Summary: [Detailed profile summary]
    Work Experience: [Work Experience details]
    Education: [Education details]
    Technical Skills: [List of technical skills]
    Projects: [Project details]
    Certificates: [Certificate details]

    Resume Text: {resume_text}
    """
    raw_extracted_text = groq_generate(prompt)
    parsed_data = parse_llm_response(raw_extracted_text)
    # Fallback extraction if necessary.
    if not parsed_data["Basic Info"]["Email"]:
        parsed_data["Basic Info"]["Email"] = extract_email(resume_text)
    if not parsed_data["Basic Info"]["Phone"]:
        parsed_data["Basic Info"]["Phone"] = extract_phone(resume_text)
    return parsed_data

def extract_linkedin(text: str) -> str:
    """Extract a LinkedIn URL from the provided text using regex."""
    pattern = r'https?://(?:www\.)?linkedin\.com/[^\s]+'
    match = re.search(pattern, text)
    return match.group(0) if match else ""

def display_section_content(section: str, data: Dict[str, Any]):
    """Display a chosen section from the resume breakdown."""
    if section not in data or not data[section]:
        st.write(f"No information available for {section}")
        return

    if section == "Basic Info":
        st.write("**Basic Information:**")
        for field in ["Name", "Email", "Phone"]:
            if data["Basic Info"].get(field):
                st.write(f"**{field}:** {data['Basic Info'].get(field)}")
            else:
                st.write(f"**{field}:** Not provided")
    elif section in ["Work Experience", "Projects"]:
        st.write(f"**{section}:**")
        content = data[section]
        if isinstance(content, list):
            for item in content:
                if section == "Work Experience":
                    st.write(f"\n**{item.get('company', '')}**")
                    st.write(f"*{item.get('duration', '')} | {item.get('location', '')}*")
                    for resp in item.get('responsibilities', []):
                        st.write(f"• {resp}")
                elif section == "Projects":
                    st.write(f"\n**{item.get('title', '')}**")
                    for detail in item.get('details', []):
                        st.write(f"• {detail}")
        else:
            st.write(content)
    else:
        st.write(f"**{section}:**")
        st.write(data[section])

def evaluate_resume_for_job(resume_text: str, job_req_text: str, threshold: float = 0.1) -> bool:
    """
    Evaluate whether a resume matches the job requirements using a simple keyword match.
    (Replace this using your Groq/LLM-based logic if desired.)
    """
    job_keywords = job_req_text.lower().split()
    resume_lower = resume_text.lower()
    score = sum(1 for kw in job_keywords if kw in resume_lower)
    if score >= len(job_keywords) * threshold:
        return True
    return False

def send_email(file_path: str, sender_email: str, sender_password: str, recipient_email: str):
    """Send the provided Excel file via email to the HR representative."""
    msg = EmailMessage()
    msg['Subject'] = "Selected Resumes for the Job"
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg.set_content("Hi,\n\nPlease find attached the Excel file containing the selected resumes based on the job requirements.\n\nBest,\nAI Screening System")
    
    with open(file_path, "rb") as f:
        file_data = f.read()
        file_name = file_path.split("/")[-1]
    msg.add_attachment(file_data, maintype="application", 
                       subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                       filename=file_name)
    
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(sender_email, sender_password)
        smtp.send_message(msg)

# ----------------------------------------------
# Main Application
# ----------------------------------------------

st.set_page_config(page_title="📄Skilltrack", layout="wide")

def main():
    mode = st.sidebar.selectbox("Select Mode", ["Individual Analysis", "Resume Screening"])
    
    if mode == "Individual Analysis":
        if "qa_system" not in st.session_state:
            st.session_state.qa_system = ResumeQASystem()
        if "resume_text" not in st.session_state:
            st.session_state.resume_text = None
        if "extracted_data" not in st.session_state:
            st.session_state.extracted_data = None

        with st.sidebar:
            st.header("📂 Upload Resume")
            uploaded_file = st.file_uploader("Choose a resume", type=["pdf", "txt"], key="single_resume")
            if uploaded_file:
                st.session_state.resume_text = read_resume(uploaded_file)
                st.session_state.extracted_data = extract_info(st.session_state.resume_text)
                st.session_state.qa_system.create_knowledge_base(st.session_state.resume_text)
                st.success("✅ Resume processed successfully!")

        st.header("📄Skilltrack-AI Resume Screening")
        tab1, tab2 = st.tabs(["🤖 AI-Powered Analysis", "📜 Resume Breakdown"])

        with tab2:
            if st.session_state.extracted_data:
                sections = ["Basic Info", "Profile Summary", "Work Experience", "Education", 
                            "Technical Skills", "Projects", "Certificates"]
                selected_section = st.selectbox("📌 Select Section:", sections, key="section_select")
                display_section_content(selected_section, st.session_state.extracted_data)
            else:
                st.warning("⚠️ Please upload a resume first.")

        with tab1:
            if st.session_state.resume_text:
                st.markdown(
                    """
                    🤖 Ask AI anything about skills, experience, projects, and qualifications.  
                    🏆 Get precise answers tailored for hiring decisions.
                    """
                )
                question = st.text_input("💬 Ask a Question:", key="question_input", help="Type your question and click 'Get Answer'")
                with st.form("qa_form", clear_on_submit=False):
                    submitted = st.form_submit_button("🚀 Get Answer", use_container_width=True)
                if submitted and question:
                    with st.spinner("🤖 Thinking..."):
                        answer = st.session_state.qa_system.answer_question(question)
                    st.markdown(
                        f"""<div style="
                            border-left: 4px solid #4CAF50;
                            padding: 12px;
                            border-radius: 8px;
                            background-color: #1e1e1e;
                            color: white;
                            margin-top: 10px;
                            font-size: 16px;
                            line-height: 1.6;">
                            <b>📄 Answer:</b><br> {answer}</div>""",
                        unsafe_allow_html=True,
                    )
                    st.markdown(
                        """
                        <style>
                            .stTextInput input { font-size: 16px; padding: 10px; border-radius: 10px; }
                            .custom-button { padding: 8px 20px; font-size: 14px; border-radius: 6px; background-color: #e74c3c; color: white; border: none; cursor: pointer; }
                            .custom-button:hover { background-color: #c0392b; }
                            .stMarkdown { font-size: 18px; line-height: 1.6; }
                        </style>
                        """,
                        unsafe_allow_html=True,
                    )
            else:
                st.warning("⚠️ Please upload a resume first.")
    
    else:
        st.header("📄Skilltrack-AI Resume Screening")
        st.sidebar.header("📂 Resume Screening")
        uploaded_resumes = st.sidebar.file_uploader("Upload Resumes (PDF or TXT)", type=["pdf", "txt"],
                                                    accept_multiple_files=True, key="bulk_resumes")
        job_req_file = st.sidebar.file_uploader("Upload Job Requirements File (PDF or TXT)", type=["pdf", "txt"],
                                                key="job_requirements")
        
        if uploaded_resumes and job_req_file:
            job_req_text = read_resume(job_req_file)
            st.info("Job requirements file uploaded and processed.")
            selected_resumes = []
            
            for file in uploaded_resumes:
                resume_text = read_resume(file)
                if evaluate_resume_for_job(resume_text, job_req_text):
                    parsed_data = extract_info(resume_text)
                    basic_info = parsed_data.get("Basic Info", {})
                    name = basic_info.get("Name", "").strip() or "NA"
                    email = basic_info.get("Email", "").strip() or "NA"
                    contact = basic_info.get("Phone", "").strip() or "NA"
                    linkedin = extract_linkedin(resume_text).strip() or "NA"
                    selected_resumes.append({
                        "Filename": file.name.strip() if file.name and file.name.strip() else "NA",
                        "Name": name,
                        "Contact": contact,
                        "Email": email,
                        "LinkedIn": linkedin
                    })

            st.write(f"**Selected Resumes:** {len(selected_resumes)} out of {len(uploaded_resumes)}")

            if selected_resumes:
                df = pd.DataFrame(selected_resumes)
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    excel_path = tmp.name
                    df.to_excel(excel_path, index=False)
                
                # Auto-adjust column widths in the generated Excel file using openpyxl.
                import openpyxl #type:ignore
                from openpyxl.utils import get_column_letter#type:ignore
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active
                for i, col in enumerate(ws.columns, start=1):
                    max_length = 0
                    col_letter = get_column_letter(i)
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_length + 2
                wb.save(excel_path)

                with open(excel_path, "rb") as f:
                    st.download_button(
                        label="Download Selected Resumes Excel",
                        data=f,
                        file_name="selected_resumes.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                st.subheader("Email Excel to HR")
                sender_email = st.text_input("Sender Email", key="sender_email")
                sender_password = st.text_input("Sender Email Password", type="password", key="sender_password")
                recipient_email = st.text_input("HR Email Address", key="recipient_email")
                if st.button("Send Email to HR"):
                    if sender_email and sender_password and recipient_email:
                        try:
                            send_email(excel_path, sender_email, sender_password, recipient_email)
                            st.success("Email sent successfully!")
                        except Exception as e:
                            st.error(f"Failed to send email: {e}")
                    else:
                        st.warning("Please provide all email details to proceed.")
            else:
                st.warning("No resumes match the job requirements. Please adjust your criteria or check the uploads.")
        else:
            st.info("Please upload both resumes and a job requirements file to proceed.")

if __name__ == "__main__":
    main()
